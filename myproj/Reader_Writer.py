import os

import openpyxl as xl
from datetime import date 


def log_in_excel(timesheet_date, master_project, project, features, ticketNo, group, activity, hours, isSaved, isSubmitted) :
    # path = os.getcwd()
    sheet_name = 'Timesheet'
    report_path = os.path.join(os.getcwd(), "Report.xlsx")
    print(report_path)

    report = xl.load_workbook(report_path)
    report_sheet = report[sheet_name]
    
    row = report_sheet.max_row

    report_sheet.cell(column=1, row=row+1, value=row)
    report_sheet.cell(column=2, row=row+1, value=date.today().strftime("%m/%d/%Y"))
    report_sheet.cell(column=3, row=row+1, value=timesheet_date)
    report_sheet.cell(column=4, row=row+1, value=master_project)
    report_sheet.cell(column=5, row=row+1, value=project)
    report_sheet.cell(column=6, row=row+1, value=features)
    report_sheet.cell(column=7, row=row+1, value=ticketNo)
    report_sheet.cell(column=8, row=row+1, value=group)
    report_sheet.cell(column=9, row=row+1, value=activity)
    report_sheet.cell(column=10, row=row+1, value=hours)
    report_sheet.cell(column=11, row=row+1, value=isSaved)
    report_sheet.cell(column=12, row=row+1, value=isSubmitted)




    report.save(report_path)
    return 1
